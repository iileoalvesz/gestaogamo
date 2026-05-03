"""
Generates an Excel file in the same layout as Sensores.xlsx:
  - One sheet per chamber (sensor.name), sensors sorted by syos_nickname
  - If chamber has 2 sensors: Início (col 1) + Fim (col 11) side by side
  - Summary block (min/max/avg/total/ok) to the right of each sensor block
  - "Resumo" sheet with overall compliance per sensor
"""

import io
from collections import defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import SensorLocation, SensorReading

# ── Palette ────────────────────────────────────────────────────────────────────
_C_PURPLE   = "6366F1"   # header / title
_C_PURPLE_L = "EEF2FF"
_C_TEAL     = "0D9488"
_C_TEAL_L   = "F0FDFA"
_C_RED      = "DC2626"
_C_RED_L    = "FEF2F2"
_C_GRAY     = "F3F4F6"
_C_HEADER   = "E5E7EB"
_C_WHITE    = "FFFFFF"
_C_BLACK    = "111827"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=_C_BLACK, size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _align(h="left", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# ── Public entry point ─────────────────────────────────────────────────────────

def generate_sensor_excel(period_start=None, period_end=None) -> bytes:
    """Returns xlsx bytes mirroring the Sensores.xlsx template."""

    sensors = list(
        SensorLocation.objects.filter(is_active=True).order_by("name", "syos_nickname")
    )

    # Group sensors by chamber name (same name = same tab)
    chambers: dict[str, list[SensorLocation]] = defaultdict(list)
    for s in sensors:
        chambers[s.name].append(s)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    resumo_rows = []  # (sensor, total, ok_count)

    for chamber_name, chamber_sensors in chambers.items():
        # Fetch readings per sensor
        readings_by_id: dict[int, list] = {}
        has_data = False
        for sensor in chamber_sensors:
            qs = SensorReading.objects.filter(sensor=sensor).order_by("recorded_at")
            if period_start:
                qs = qs.filter(recorded_at__date__gte=period_start)
            if period_end:
                qs = qs.filter(recorded_at__date__lte=period_end)
            rds = list(qs)
            readings_by_id[sensor.id] = rds
            if rds:
                has_data = True

        if not has_data:
            continue

        ws = wb.create_sheet(title=_safe_name(chamber_name))

        # Up to 2 sensors per sheet (Início = first, Fim = second)
        active = [(s, readings_by_id[s.id]) for s in chamber_sensors if readings_by_id[s.id]]
        labels = ["Inicio", "Fim"]
        col_starts = [1, 11]   # data block column start (1-based)

        for idx, (sensor, readings) in enumerate(active[:2]):
            _write_block(ws, sensor, readings, col_starts[idx], labels[idx])
            total = len(readings)
            ok_count = sum(1 for r in readings if r.is_ok)
            resumo_rows.append((sensor, total, ok_count))

    _write_resumo(wb, resumo_rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sheet block writer ─────────────────────────────────────────────────────────

def _write_block(ws, sensor: SensorLocation, readings: list, col_start: int, label: str):
    min_t = float(sensor.min_temp) if sensor.min_temp is not None else None
    max_t = float(sensor.max_temp) if sensor.max_temp is not None else None
    std   = sensor.temp_standard or ""

    # ── Row 1: title ────────────────────────────────────────────────────────
    title = f"{sensor.name} - {label} - (Padrão {std}) - {sensor.syos_nickname}"
    tc = ws.cell(row=1, column=col_start, value=title)
    tc.font      = _font(bold=True, color=_C_WHITE, size=11)
    tc.fill      = _fill(_C_PURPLE)
    tc.alignment = _align("center")
    ws.row_dimensions[1].height = 22
    ws.merge_cells(
        start_row=1, start_column=col_start,
        end_row=1,   end_column=col_start + 6,
    )

    # ── Row 2: column headers ────────────────────────────────────────────────
    min_disp = int(min_t) if min_t is not None and min_t == int(min_t) else (min_t or "?")
    max_disp = int(max_t) if max_t is not None and max_t == int(max_t) else (max_t or "?")
    headers = [
        "N° REGISTROS", "DATA", "HORA", "TEMPERATURA °C",
        f"META MÍNIMA ({min_disp}°C)", f"META MÁXIMA ({max_disp}°C)", "OK/NÃO OK",
    ]
    for i, h in enumerate(headers):
        c = ws.cell(row=2, column=col_start + i, value=h)
        c.font      = _font(bold=True)
        c.fill      = _fill(_C_HEADER)
        c.alignment = _align("center", wrap=True)
    ws.row_dimensions[2].height = 28

    # ── Rows 3+: data rows ───────────────────────────────────────────────────
    temps = [float(r.temperature) for r in readings]
    for idx, reading in enumerate(readings):
        row   = 3 + idx
        dt    = reading.recorded_at
        temp  = float(reading.temperature)
        is_ok = reading.is_ok

        row_bg = _fill(_C_RED_L) if not is_ok else PatternFill()

        values = [
            (idx + 1,                       "center"),
            (dt.date(),                      "left"),
            (dt.strftime("%H:%M"),           "center"),
            (round(temp, 2),                 "right"),
            (min_t,                          "center"),
            (max_t,                          "center"),
            ("OK" if is_ok else "NÃO OK",   "center"),
        ]
        for ci, (val, align) in enumerate(values):
            c = ws.cell(row=row, column=col_start + ci, value=val)
            c.alignment = _align(align)
            if not is_ok:
                c.fill = row_bg
            if ci == 3:  # temperature
                c.font = _font(bold=True, color=_C_RED if not is_ok else _C_BLACK)
            elif ci == 6:  # ok label
                c.font = _font(bold=True, color=_C_TEAL if is_ok else _C_RED)
            else:
                c.font = _font()
            if ci == 1:  # date: format
                c.number_format = "DD/MM/YYYY"

    # ── Summary block (rows 3–7, right of data) ──────────────────────────────
    sum_col = col_start + 7   # one gap column after data block
    if readings:
        avg_t = sum(temps) / len(temps)
        ok_count = sum(1 for r in readings if r.is_ok)
        summary = [
            ("TEMPERATURA MÍNIMA °C", round(min(temps), 2)),
            ("TEMPERATURA MÁXIMA °C", round(max(temps), 2)),
            ("TEMPERATURA MÉDIA °C",  round(avg_t, 2)),
            ("TOTAL REGISTROS",       len(readings)),
            ("TOTAL REGISTROS OK",    ok_count),
        ]
        for i, (lbl, val) in enumerate(summary):
            r = 3 + i
            is_ok_row = "OK" in lbl
            lc = ws.cell(row=r, column=sum_col, value=lbl)
            lc.font      = _font(bold=True)
            lc.fill      = _fill(_C_TEAL_L if is_ok_row else _C_GRAY)
            lc.alignment = _align("left")

            vc = ws.cell(row=r, column=sum_col + 1, value=val)
            vc.font      = _font(bold=True, color=_C_TEAL if is_ok_row else _C_BLACK)
            vc.fill      = _fill(_C_TEAL_L if is_ok_row else _C_GRAY)
            vc.alignment = _align("center")

    # ── Column widths ────────────────────────────────────────────────────────
    widths = [13, 13, 9, 16, 20, 20, 12, 26, 13]
    for i, w in enumerate(widths):
        col_letter = get_column_letter(col_start + i)
        cur = ws.column_dimensions[col_letter].width or 0
        ws.column_dimensions[col_letter].width = max(cur, w)


# ── Resumo sheet ───────────────────────────────────────────────────────────────

def _write_resumo(wb: openpyxl.Workbook, summary_rows: list):
    ws = wb.create_sheet(title="Resumo")

    # Row 1: title
    ws.merge_cells("A1:F1")
    tc = ws.cell(1, 1, "% REGISTROS DE TEMPERATURAS CONFORMES, T°C MÍNIMA E MÁXIMA")
    tc.font      = _font(bold=True, color=_C_WHITE, size=12)
    tc.fill      = _fill(_C_PURPLE)
    tc.alignment = _align("center")
    ws.row_dimensions[1].height = 26

    # Row 2: headers
    hdrs = [
        "Local", "T°C padrão ambientes", "N° registros",
        "N° registros conformes", "% registros T°C CONFORME SF", "Meta",
    ]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(2, i, h)
        c.font      = _font(bold=True)
        c.fill      = _fill(_C_HEADER)
        c.alignment = _align("center", wrap=True)
    ws.row_dimensions[2].height = 32

    META = 0.85
    for row_idx, (sensor, total, ok_count) in enumerate(summary_rows, 3):
        pct = round(ok_count / total, 4) if total else 0
        ok  = pct >= META

        row_vals = [sensor.name, sensor.temp_standard or "—", total, ok_count, pct, META]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row_idx, ci, val)
            c.font      = _font()
            c.alignment = _align("center" if ci >= 3 else "left")
            if ci == 5:
                c.number_format = "0.0%"
                c.font = _font(bold=True, color=_C_TEAL if ok else _C_RED)
                c.fill = _fill(_C_TEAL_L if ok else _C_RED_L)
            elif ci == 6:
                c.number_format = "0%"
                c.fill = _fill(_C_GRAY)

    # Column widths
    for col, w in zip("ABCDEF", [38, 22, 14, 22, 28, 10]):
        ws.column_dimensions[col].width = w


def _safe_name(name: str) -> str:
    for ch in r"\/?*[]:":
        name = name.replace(ch, "-")
    return name[:31]
