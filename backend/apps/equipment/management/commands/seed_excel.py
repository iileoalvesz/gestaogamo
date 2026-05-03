"""
Management command para importar dados da planilha Excel de controle de calibração.

Uso:
    python manage.py seed_excel --file "caminho/para/planilha.xlsx"
"""
import re
from datetime import datetime, date, timedelta
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from apps.equipment.models import Equipment

SHEET_TYPE_MAP = {
    "BALANÇAS": "balanca",
    "BALANÇAS ": "balanca",
    "SYOS": "syos",
    "LOGGER": "logger",
    "PAQUIMETRO": "paquimetro",
    "PAQUÍMETRO": "paquimetro",
    "TERMOMETRO ESPETO_2025": "termometro",
    "TERMÔMETRO ESPETO_2025": "termometro",
}

SKIP_SHEETS = {"CARJ", "Identificação", "Extraviado", "Extravios", "Emails"}

SITUATION_MAP = {
    "calibrado": "ativo",
    "ativo": "ativo",
    "danificado": "danificado",
    "perdido": "perdido",
    "fora de uso": "fora_de_uso",
    "fora_de_uso": "fora_de_uso",
    "extraviado": "perdido",
}

HEADER_ALIASES = {
    "equipamento": "name",
    "local": "location",
    "modelo": "model",
    "identificação": "identification",
    "identificacao": "identification",
    "n.º série": "serial_number",
    "n.o serie": "serial_number",
    "no serie": "serial_number",
    "n° série": "serial_number",
    "série": "serial_number",
    "serie": "serial_number",
    "certificado (n.º)": "certificate_number",
    "certificado (n.o)": "certificate_number",
    "certificado n.o": "certificate_number",
    "n° certificado": "certificate_number",
    "n.º certificado": "certificate_number",
    "frequência de calibração": "calibration_frequency",
    "frequencia de calibracao": "calibration_frequency",
    "data da última calibração": "last_calibration_date",
    "data da ultima calibracao": "last_calibration_date",
    "data da proxima calibração": "next_calibration_date",
    "data da proxima calibracao": "next_calibration_date",
    "data da próxima calibração": "next_calibration_date",
    "status": "status_col",
    "situação": "situation",
    "situacao": "situation",
    "certificado": "certificate_file_name",
}


def _normalize_header(h):
    if h is None:
        return ""
    return str(h).strip().lower().replace("\n", " ")


def _parse_date(value):
    if value is None:
        return None
    if isinstance(value, (datetime,)):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _parse_frequency(value):
    if value is None:
        return 365
    if isinstance(value, (int, float)):
        return int(value)
    raw = str(value).strip().lower()
    if "anual" in raw or "annual" in raw:
        return 365
    if "semestral" in raw:
        return 180
    if "trimestral" in raw:
        return 90
    if "mensal" in raw:
        return 30
    digits = re.findall(r"\d+", raw)
    if digits:
        return int(digits[0])
    return 365


def _parse_situation(value):
    if value is None:
        return "ativo"
    raw = str(value).strip().lower()
    return SITUATION_MAP.get(raw, "ativo")


def _find_header_row(sheet, max_rows=5):
    """Busca a linha cujo primeiro valor não nulo seja exatamente 'Equipamento'."""
    for row_idx in range(1, max_rows + 1):
        for cell in sheet[row_idx]:
            if cell.value is None:
                continue
            val = str(cell.value).strip().lower()
            # Row title contains "equipamentos" (plural) — skip
            if val == "equipamento":
                return row_idx
            break  # first non-null cell per row; stop if it's not 'equipamento'
    return None


def _map_columns(sheet, header_row):
    """Retorna dict {field_name: col_index} baseado nos cabeçalhos da planilha."""
    mapping = {}
    for cell in sheet[header_row]:
        if cell.value is None:
            continue
        normalized = _normalize_header(cell.value)
        field = HEADER_ALIASES.get(normalized)
        if field:
            mapping[field] = cell.column - 1  # 0-indexed
    return mapping


class Command(BaseCommand):
    help = "Importa dados de equipamentos da planilha Excel de controle de calibração"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            default=None,
            help="Caminho para o arquivo .xlsx",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Remove todos os equipamentos antes de importar",
        )

    def handle(self, *args, **options):
        file_path = options.get("file")

        if not file_path:
            # Default: procura na pasta pai do projeto
            base = Path(__file__).resolve().parents[6]
            candidates = list(base.glob("**/Formulario Controle de Calibracao*.xlsx"))
            if not candidates:
                raise CommandError(
                    "Arquivo Excel não encontrado. Use --file para especificar o caminho."
                )
            file_path = str(candidates[0])
            self.stdout.write(f"Usando arquivo: {file_path}")

        if not Path(file_path).exists():
            raise CommandError(f"Arquivo não encontrado: {file_path}")

        if options["clear"]:
            count = Equipment.objects.count()
            Equipment.objects.all().delete()
            self.stdout.write(self.style.WARNING(f"Removidos {count} equipamentos existentes."))

        wb = load_workbook(filename=file_path, data_only=True)
        total_imported = 0
        total_updated = 0

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue

            eq_type = SHEET_TYPE_MAP.get(sheet_name)
            if eq_type is None:
                # Try partial match
                for key, val in SHEET_TYPE_MAP.items():
                    if key.upper() in sheet_name.upper() or sheet_name.upper() in key.upper():
                        eq_type = val
                        break
                if eq_type is None:
                    self.stdout.write(
                        self.style.WARNING(f"  Aba '{sheet_name}' ignorada (tipo não reconhecido).")
                    )
                    continue

            sheet = wb[sheet_name]
            header_row = _find_header_row(sheet)
            if header_row is None:
                self.stdout.write(
                    self.style.WARNING(f"  Aba '{sheet_name}': cabeçalho não encontrado, ignorando.")
                )
                continue

            col_map = _map_columns(sheet, header_row)
            if "name" not in col_map:
                self.stdout.write(
                    self.style.WARNING(f"  Aba '{sheet_name}': coluna 'Equipamento' não mapeada, ignorando.")
                )
                continue

            sheet_imported = 0
            sheet_updated = 0

            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                def col(field):
                    idx = col_map.get(field)
                    if idx is not None and idx < len(row):
                        return row[idx]
                    return None

                name = col("name")
                if not name or str(name).strip() == "":
                    continue

                name = str(name).strip()
                identification = str(col("identification") or "").strip()
                serial_number = str(col("serial_number") or "").strip()
                certificate_number = str(col("certificate_number") or "").strip()
                location = str(col("location") or "").strip()
                model = str(col("model") or "").strip()
                last_cal = _parse_date(col("last_calibration_date"))
                next_cal = _parse_date(col("next_calibration_date"))
                freq = _parse_frequency(col("calibration_frequency"))
                situation = _parse_situation(col("situation"))

                # Se next_cal não veio da planilha mas temos last_cal, calcula
                if last_cal and not next_cal:
                    next_cal = last_cal + timedelta(days=freq)

                defaults = {
                    "name": name,
                    "equipment_type": eq_type,
                    "location": location,
                    "model": model,
                    "serial_number": serial_number,
                    "certificate_number": certificate_number,
                    "calibration_frequency_days": freq,
                    "last_calibration_date": last_cal,
                    "next_calibration_date": next_cal,
                    "situation": situation,
                }

                lookup = {"name": name, "equipment_type": eq_type}
                if identification:
                    lookup = {"identification": identification, "equipment_type": eq_type}

                obj, created = Equipment.objects.update_or_create(
                    **lookup, defaults=defaults
                )
                if not obj.identification and identification:
                    obj.identification = identification
                    obj.save(update_fields=["identification"])

                if created:
                    sheet_imported += 1
                else:
                    sheet_updated += 1

            self.stdout.write(
                f"  Aba '{sheet_name}' ({eq_type}): "
                f"{sheet_imported} importados, {sheet_updated} atualizados."
            )
            total_imported += sheet_imported
            total_updated += sheet_updated

        self.stdout.write(
            self.style.SUCCESS(
                f"\nTotal: {total_imported} equipamentos importados, {total_updated} atualizados."
            )
        )
